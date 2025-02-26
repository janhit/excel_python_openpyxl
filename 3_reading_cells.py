import openpyxl

wb = openpyxl.Workbook()

ws = wb.worksheets[0]

ws["A1"].value = 56
ws["C3"].value = 123.0
val_1 = ws.cell(1+2,2+1).value

print(type(val_1))


