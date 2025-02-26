import openpyxl

wb = openpyxl.Workbook()
wb.save("first_wb.xlsx")

wb = openpyxl.load_workbook("test.xlsx")

ws_1 = wb.create_sheet("A Sheet-1", 0)
ws_2 = wb.create_sheet("A Sheet-2")

#wb.remove(wb["Sheet"])

#del wb["Tabelle1"]

for sheet in wb:
    print(sheet.title)

ws_2.title = "New Title"

wb.save("test.xlsx")

